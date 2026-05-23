import marimo

__generated_with = "0.23.5"
app = marimo.App(width="medium")


@app.cell(hide_code=True)
def _():
    import marimo as mo

    return (mo,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    # Flow 01 · Ingestion

    Extracts OSM infrastructure layers from a local PBF file into Bronze GeoParquet,
    then downloads EIA Form 860/923 and EPA FRS data.

    **Interactive:** fill in the form below and click **Run Ingestion**.
    **Script:** `marimo run flows/01_ingest.py -- --help` (or `-- --config ... --run-osm true`)
    """)
    return


@app.cell
def _():
    import time
    import zipfile
    from pathlib import Path

    import httpx
    from tqdm import tqdm
    from pydantic import BaseModel, Field

    import sys as _sys
    _sys.path.insert(0, str(Path(".").resolve()))
    from src.lifelinepoi.config import LifelineConfig
    from lib.duckdb_utils import get_connection, run_layer_sql

    return (
        BaseModel,
        Field,
        LifelineConfig,
        get_connection,
        httpx,
        run_layer_sql,
        tqdm,
        zipfile,
    )


@app.cell
def _(BaseModel, Field):
    class FlowParams(BaseModel):
        config_path: str = Field(default="config.lifeline.yaml", description="Path to config YAML file")
        run_osm: bool = Field(default=True, description="Run OSM PBF extraction")
        run_eia: bool = Field(default=True, description="Download EIA Form 860/923")
        run_epa: bool = Field(default=True, description="Download EPA FRS data")
        run_echo: bool = Field(default=True, description="Download EPA ECHO facility export")
        run_sdwis: bool = Field(default=True, description="Download EPA SDWIS drinking water data")
        run_fcc: bool = Field(default=True, description="Download FCC ASR cell tower/antenna data")
        run_irs_tcn: bool = Field(default=True, description="Download IRS fuel terminal control numbers")
        run_eia_terminals: bool = Field(default=True, description="Download EIA terminal data (requires EIA_API_KEY in .env.local)")
        run_hifld: bool = Field(default=True, description="Download HIFLD validation layers from seerai-hifld-archive GCS")
        run_cms: bool = Field(default=True, description="Download CMS Hospital & Non-Hospital Provider data")
        area: str = Field(default="", description="Area name to process (empty = all configured areas)")
        layer: str = Field(default="", description="Single OSM layer (empty = all layers)")

    return (FlowParams,)


@app.cell
def _(mo):
    params_form = (
        mo.md("""
        ## Parameters

        Config file: {config_path}

        Steps to run:
        - {run_osm} OSM PBF extraction
        - {run_eia} EIA Form 860/923 download
        - {run_epa} EPA FRS download
        - {run_echo} EPA ECHO facility export
        - {run_sdwis} EPA SDWIS drinking water
        - {run_fcc} FCC ASR cell towers
        - {run_irs_tcn} IRS fuel terminal control numbers
        - {run_eia_terminals} EIA terminal locations (needs `EIA_API_KEY` in `.env.local`)
        - {run_hifld} HIFLD validation layers (seerai-hifld-archive)
        - {run_cms} CMS hospital provider data (bed counts)

        Single layer override (blank = all layers): {layer}

        Area override (blank = all configured areas): {area}
        """)
        .batch(
            config_path=mo.ui.text(value="config.lifeline.yaml", label="Config path"),
            run_osm=mo.ui.checkbox(value=True, label=""),
            run_eia=mo.ui.checkbox(value=True, label=""),
            run_epa=mo.ui.checkbox(value=True, label=""),
            run_echo=mo.ui.checkbox(value=True, label=""),
            run_sdwis=mo.ui.checkbox(value=True, label=""),
            run_fcc=mo.ui.checkbox(value=True, label=""),
            run_irs_tcn=mo.ui.checkbox(value=True, label=""),
            run_eia_terminals=mo.ui.checkbox(value=True, label=""),
            run_hifld=mo.ui.checkbox(value=True, label=""),
            run_cms=mo.ui.checkbox(value=True, label=""),
            layer=mo.ui.text(value="", placeholder="e.g. power", label=""),
            area=mo.ui.text(value="", placeholder="e.g. florida", label=""),
        )
        .form(submit_button_label="▶ Run Ingestion")
    )
    params_form
    return (params_form,)


@app.cell
def _(FlowParams, mo):
    import sys as _sys
    is_script_mode = mo.app_meta().mode == "script"
    if is_script_mode and "help" in mo.cli_args():
        print("Usage: marimo run flows/01_ingest.py -- [options]\n")
        for _name, _field in FlowParams.model_fields.items():
            _default = f"(default: {_field.default})" if _field.default is not None else "(required)"
            print(f"  --{_name.replace('_', '-'):<28} {_field.description} {_default}")
        _sys.exit(0)
    return (is_script_mode,)


@app.cell
def _(FlowParams, is_script_mode, mo, params_form):
    mo.stop(
        not is_script_mode and params_form.value is None,
        mo.callout(mo.md("**Fill in the parameters above and click _Run Ingestion_ to start.**"), kind="info"),
    )
    if is_script_mode:
        flow_params = FlowParams(**{k.replace("-", "_"): v for k, v in mo.cli_args().items()})
    else:
        flow_params = FlowParams(**params_form.value)
    return (flow_params,)


@app.cell
def _(LifelineConfig, flow_params, mo):
    cfg = LifelineConfig.from_yaml(flow_params.config_path)
    _area_names = [a.name for a in cfg.areas]
    mo.md(f"**Config loaded.** Layers: `{', '.join(cfg.osm.layers)}`  |  Areas: `{', '.join(_area_names)}`")
    return (cfg,)


@app.cell
def _(cfg, flow_params, get_connection, mo, run_layer_sql):
    def _ingest_osm_area(area, layers=None):
        from pathlib import Path as _P
        import time as _t
        pbf_path = _P(area.pbf_path)
        if not pbf_path.exists():
            raise FileNotFoundError(
                f"OSM PBF file not found for area '{area.name}': {pbf_path}\n"
                "Download a regional extract from https://download.geofabrik.de/ "
                "and update the pbf_path in config.lifeline.yaml"
            )
        sql_dir = _P(".").resolve() / "sql"
        bronze_osm = _P(cfg.storage.bronze_path) / "osm" / area.name
        bronze_osm.mkdir(parents=True, exist_ok=True)
        target_layers = layers or cfg.osm.layers
        print(f"[{area.name}] OSM PBF: {pbf_path}")
        print(f"[{area.name}] Layers to extract: {target_layers}")
        conn = get_connection(cfg.duckdb.memory_limit)
        try:
            for _lyr in target_layers:
                output_path = bronze_osm / f"{_lyr}.parquet"
                print(f"  [{area.name}] Extracting layer: {_lyr} -> {output_path}")
                t0 = _t.perf_counter()
                run_layer_sql(
                    conn=conn, sql_dir=sql_dir, layer=_lyr,
                    pbf_path=str(pbf_path), output_path=str(output_path),
                    osmium_index_type=cfg.osm.osmium_index_type,
                )
                elapsed = _t.perf_counter() - t0
                size_mb = output_path.stat().st_size / 1_048_576
                print(f"    [{area.name}] Done in {elapsed:.1f}s — {size_mb:.1f} MB")
        finally:
            conn.close()

    if flow_params.run_osm:
        area_filter = flow_params.area.strip()
        target_areas = [a for a in cfg.areas if not area_filter or a.name == area_filter]
        if not target_areas:
            osm_result = mo.callout(
                mo.md(f"⚠️ No area named `{area_filter}` found in config. Available: `{', '.join(a.name for a in cfg.areas)}`"),
                kind="warn",
            )
        else:
            layer_list = [flow_params.layer] if flow_params.layer else None
            print(f"[OSM] Processing {len(target_areas)} area(s): {', '.join(a.name for a in target_areas)}")
            for _area in target_areas:
                _ingest_osm_area(_area, layers=layer_list)
            osm_result = mo.callout(
                mo.md(f"✅ **OSM extraction complete.** Areas: `{', '.join(a.name for a in target_areas)}`"),
                kind="success",
            )
    else:
        osm_result = mo.callout(mo.md("⏭ OSM extraction skipped."), kind="neutral")
    osm_result
    return (osm_result,)


@app.cell
def _(cfg, flow_params, httpx, mo, tqdm, zipfile):
    def _download_file(url, dest, label):
        from pathlib import Path as _P
        dest.parent.mkdir(parents=True, exist_ok=True)
        print(f"  Downloading {label} from {url}")
        with httpx.stream("GET", url, follow_redirects=True, timeout=120) as r:
            r.raise_for_status()
            total = int(r.headers.get("content-length", 0))
            with open(dest, "wb") as f, tqdm(
                total=total, unit="B", unit_scale=True, desc=label, leave=False
            ) as bar:
                for chunk in r.iter_bytes(chunk_size=65536):
                    f.write(chunk)
                    bar.update(len(chunk))
        return dest

    def _extract_zip(zip_path, extract_dir):
        """Extract zip to extract_dir, skip if sentinel file already present."""
        from pathlib import Path as _P
        sentinel = extract_dir / ".extracted"
        if sentinel.exists():
            print(f"    Already extracted → {extract_dir}")
            return
        extract_dir.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(zip_path) as z:
            members = z.namelist()
            print(f"    Extracting {len(members)} files → {extract_dir}")
            z.extractall(extract_dir)
        sentinel.touch()
        print(f"    Extraction complete.")

    def _eia860_plants_to_geoparquet(eia860_dir):
        """Convert 2___Plant_Y*.xlsx → eia860_plants.parquet (GeoParquet)."""
        import duckdb as _duckdb
        from pathlib import Path as _P
        # Find the plant file (name changes with year)
        candidates = sorted(_P(eia860_dir).glob("2___Plant_Y*.xlsx"))
        if not candidates:
            print("  EIA 860: no Plant XLSX found, skipping GeoParquet conversion.")
            return
        plant_xlsx = candidates[-1]  # latest year if multiple
        parquet_out = _P(eia860_dir).parent / "eia860_plants.parquet"
        if parquet_out.exists():
            print(f"  EIA 860 plants GeoParquet: already exists → {parquet_out}")
            return
        xlsx_path = str(plant_xlsx).replace("\\", "/")
        out_path = str(parquet_out).replace("\\", "/")
        print(f"  Converting {plant_xlsx.name} → {parquet_out.name}")
        conn = _duckdb.connect()
        try:
            conn.execute("INSTALL spatial; LOAD spatial;")
            conn.execute(f"""
                COPY (
                    SELECT
                        * EXCLUDE ("Latitude", "Longitude"),
                        CASE
                            WHEN TRY_CAST("Latitude" AS DOUBLE) IS NOT NULL
                             AND TRY_CAST("Longitude" AS DOUBLE) IS NOT NULL
                            THEN ST_Point(
                                TRY_CAST("Longitude" AS DOUBLE),
                                TRY_CAST("Latitude"  AS DOUBLE)
                            )
                            ELSE NULL
                        END AS geometry
                    FROM read_xlsx(
                        '{xlsx_path}',
                        sheet='Plant',
                        range='A2:AP10000',
                        header=true,
                        all_varchar=true
                    )
                    WHERE "Plant Code" IS NOT NULL AND "Plant Code" != ''
                ) TO '{out_path}' (FORMAT PARQUET)
            """)
            count = conn.execute(f"SELECT COUNT(*) FROM '{out_path}'").fetchone()[0]
            print(f"  EIA 860 plants: {count:,} rows → {parquet_out.name} "
                  f"({parquet_out.stat().st_size / 1_048_576:.1f} MB)")
        finally:
            conn.close()

    def _ingest_eia():
        from pathlib import Path as _P
        bronze_eia = _P(cfg.storage.bronze_path) / "eia"
        bronze_eia.mkdir(parents=True, exist_ok=True)
        products = [
            (cfg.eia.form860_zip_url, "eia860_2023.zip", bronze_eia / "eia860"),
            (cfg.eia.form923_zip_url, "eia923_2023.zip", bronze_eia / "eia923"),
        ]
        for url, name, extract_dir in products:
            dest = bronze_eia / name
            # Download if not already on disk
            if not dest.exists():
                try:
                    _download_file(url, dest, name)
                    print(f"  EIA {name}: downloaded ({dest.stat().st_size / 1_048_576:.1f} MB)")
                except zipfile.BadZipFile:
                    dest.unlink(missing_ok=True)
                    print(f"  ERROR: {name} is not a valid zip — check eia.*_zip_url in config.")
                    continue
                except Exception as e:
                    print(f"  WARNING: Failed to download {name}: {e}")
                    continue
            else:
                print(f"  EIA {name}: already downloaded, skipping download")
            # Extract
            try:
                _extract_zip(dest, extract_dir)
            except zipfile.BadZipFile:
                print(f"  ERROR: {name} on disk is corrupt — delete it and re-run to re-download.")
        # Convert EIA 860 plant file to GeoParquet
        _eia860_plants_to_geoparquet(bronze_eia / "eia860")

    if flow_params.run_eia:
        print("[EIA] Starting Form 860/923 download + extraction")
        _ingest_eia()
        eia_result = mo.callout(mo.md("✅ **EIA download + extraction complete.**"), kind="success")
    else:
        eia_result = mo.callout(mo.md("⏭ EIA download skipped."), kind="neutral")
    eia_result
    return (eia_result,)


@app.cell
def _(cfg, flow_params, httpx, mo, tqdm, zipfile):
    def _ingest_epa():
        import duckdb as _duckdb
        from pathlib import Path as _P
        bronze_epa = _P(cfg.storage.bronze_path) / "epa"
        bronze_epa.mkdir(parents=True, exist_ok=True)
        dest = bronze_epa / "frs_national_single.zip"
        extract_dir = bronze_epa / "frs_national"
        extract_sentinel = extract_dir / ".extracted"
        parquet_out = bronze_epa / "frs_national.parquet"
        url = cfg.epa.frs_zip_url

        # --- Download ---
        if not dest.exists():
            try:
                print(f"  Downloading EPA FRS national_single from {url}")
                with httpx.stream("GET", url, follow_redirects=True, timeout=300) as r:
                    r.raise_for_status()
                    total = int(r.headers.get("content-length", 0))
                    with open(dest, "wb") as f, tqdm(
                        total=total, unit="B", unit_scale=True, desc="frs_national_single.zip", leave=False
                    ) as bar:
                        for chunk in r.iter_bytes(chunk_size=65536):
                            f.write(chunk)
                            bar.update(len(chunk))
                print(f"  EPA FRS: downloaded ({dest.stat().st_size / 1_048_576:.1f} MB)")
            except httpx.HTTPStatusError as e:
                dest.unlink(missing_ok=True)
                print(f"  ERROR: EPA FRS download failed (HTTP {e.response.status_code}). "
                      f"Update epa.frs_zip_url in config.lifeline.yaml.")
                return
            except Exception as e:
                dest.unlink(missing_ok=True)
                print(f"  WARNING: Failed to download EPA FRS: {e}")
                return
        else:
            print("  EPA FRS: already downloaded, skipping download")

        # --- Extract CSV ---
        if not extract_sentinel.exists():
            try:
                extract_dir.mkdir(parents=True, exist_ok=True)
                with zipfile.ZipFile(dest) as z:
                    members = z.namelist()
                    print(f"  Extracting {len(members)} files → {extract_dir}")
                    z.extractall(extract_dir)
                extract_sentinel.touch()
                print("  EPA FRS extraction complete.")
            except zipfile.BadZipFile:
                dest.unlink(missing_ok=True)
                print("  ERROR: frs_national_single.zip is corrupt — deleted, re-run to re-download.")
                return
        else:
            print(f"  EPA FRS: already extracted → {extract_dir}")

        # --- Convert to GeoParquet ---
        if parquet_out.exists():
            print(f"  EPA FRS GeoParquet: already exists → {parquet_out}")
            return
        csv_files = sorted(extract_dir.glob("*.csv")) + sorted(extract_dir.glob("*.CSV"))
        if not csv_files:
            print("  WARNING: No CSV files found in extract dir, skipping GeoParquet conversion.")
            return
        csv_list = "[" + ", ".join(f"'{str(p).replace(chr(92), '/')}'" for p in csv_files) + "]"
        print(f"  Converting {len(csv_files)} CSV(s) to GeoParquet → {parquet_out}")
        conn = _duckdb.connect()
        try:
            conn.execute("INSTALL spatial; LOAD spatial;")
            conn.execute(f"""
                COPY (
                    SELECT
                        * EXCLUDE (LATITUDE83, LONGITUDE83),
                        CASE
                            WHEN LATITUDE83 IS NOT NULL AND LONGITUDE83 IS NOT NULL
                            THEN ST_Transform(
                                ST_Point(LONGITUDE83, LATITUDE83),
                                'EPSG:4269', 'EPSG:4326'
                            )
                            ELSE NULL
                        END AS geometry
                    FROM read_csv(
                        {csv_list},
                        header=True,
                        union_by_name=True,
                        dateformat='%d-%b-%y',
                        ignore_errors=True
                    )
                ) TO '{str(parquet_out).replace(chr(92), '/')}' (FORMAT PARQUET)
            """)
            row_count = conn.execute(
                f"SELECT COUNT(*) FROM '{str(parquet_out).replace(chr(92), '/')}'"
            ).fetchone()[0]
            print(f"  EPA FRS GeoParquet: {row_count:,} rows written ({parquet_out.stat().st_size / 1_048_576:.1f} MB)")
        finally:
            conn.close()

    if flow_params.run_epa:
        print("[EPA] Starting FRS download + extraction + GeoParquet conversion")
        _ingest_epa()
        epa_result = mo.callout(mo.md("✅ **EPA FRS → GeoParquet complete.**"), kind="success")
    else:
        epa_result = mo.callout(mo.md("⏭ EPA download skipped."), kind="neutral")
    epa_result
    return (epa_result,)


@app.cell
def _(cfg, flow_params, httpx, mo, tqdm, zipfile):
    def _ingest_echo():
        import duckdb as _duckdb
        from pathlib import Path as _P
        bronze_epa = _P(cfg.storage.bronze_path) / "epa"
        bronze_epa.mkdir(parents=True, exist_ok=True)
        dest = bronze_epa / "echo_exporter.zip"
        extract_dir = bronze_epa / "echo"
        parquet_out = bronze_epa / "echo_exporter.parquet"
        url = cfg.echo.exporter_zip_url

        if not dest.exists():
            try:
                print(f"  Downloading EPA ECHO Exporter from {url}")
                with httpx.stream("GET", url, follow_redirects=True, timeout=600) as r:
                    r.raise_for_status()
                    total = int(r.headers.get("content-length", 0))
                    with open(dest, "wb") as f, tqdm(
                        total=total, unit="B", unit_scale=True, desc="echo_exporter.zip", leave=False
                    ) as bar:
                        for chunk in r.iter_bytes(chunk_size=65536):
                            f.write(chunk)
                            bar.update(len(chunk))
                print(f"  EPA ECHO: downloaded ({dest.stat().st_size / 1_048_576:.1f} MB)")
            except httpx.HTTPStatusError as e:
                dest.unlink(missing_ok=True)
                print(f"  ERROR: EPA ECHO download failed (HTTP {e.response.status_code})")
                return
            except Exception as e:
                dest.unlink(missing_ok=True)
                print(f"  WARNING: Failed to download EPA ECHO: {e}")
                return
        else:
            print("  EPA ECHO: already downloaded, skipping download")

        sentinel = extract_dir / ".extracted"
        if not sentinel.exists():
            extract_dir.mkdir(parents=True, exist_ok=True)
            try:
                with zipfile.ZipFile(dest) as z:
                    print(f"  Extracting {len(z.namelist())} files → {extract_dir}")
                    z.extractall(extract_dir)
                sentinel.touch()
            except zipfile.BadZipFile:
                dest.unlink(missing_ok=True)
                print("  ERROR: echo_exporter.zip is corrupt — deleted, re-run to re-download.")
                return
        else:
            print(f"  EPA ECHO: already extracted → {extract_dir}")

        if parquet_out.exists():
            print(f"  EPA ECHO GeoParquet: already exists → {parquet_out}")
            return

        csv_files = (
            sorted(extract_dir.glob("ECHO_EXPORTER*.csv"))
            + sorted(extract_dir.glob("echo_exporter*.csv"))
            + sorted(extract_dir.glob("*.csv"))
        )
        if not csv_files:
            print("  WARNING: No CSV found in ECHO extract, skipping GeoParquet conversion.")
            return

        csv_path = str(csv_files[0]).replace("\\", "/")
        out_path = str(parquet_out).replace("\\", "/")
        print(f"  Converting {csv_files[0].name} → {parquet_out.name}")
        conn = _duckdb.connect()
        try:
            conn.execute("INSTALL spatial; LOAD spatial;")
            conn.execute(f"""
                COPY (
                    SELECT
                        * EXCLUDE (FAC_LAT, FAC_LONG),
                        CASE
                            WHEN TRY_CAST(FAC_LAT AS DOUBLE) IS NOT NULL
                             AND TRY_CAST(FAC_LONG AS DOUBLE) IS NOT NULL
                            THEN ST_Point(TRY_CAST(FAC_LONG AS DOUBLE), TRY_CAST(FAC_LAT AS DOUBLE))
                            ELSE NULL
                        END AS geometry
                    FROM read_csv('{csv_path}', header=True, ignore_errors=True, all_varchar=True)
                ) TO '{out_path}' (FORMAT PARQUET)
            """)
            count = conn.execute(f"SELECT COUNT(*) FROM '{out_path}'").fetchone()[0]
            print(f"  EPA ECHO: {count:,} rows → {parquet_out.name} ({parquet_out.stat().st_size / 1_048_576:.1f} MB)")
        except Exception as e:
            print(f"  ERROR converting ECHO CSV: {e}")
            parquet_out.unlink(missing_ok=True)
        finally:
            conn.close()

    if flow_params.run_echo:
        print("[ECHO] Starting EPA ECHO Exporter download + GeoParquet conversion")
        _ingest_echo()
        echo_result = mo.callout(mo.md("✅ **EPA ECHO → GeoParquet complete.**"), kind="success")
    else:
        echo_result = mo.callout(mo.md("⏭ EPA ECHO download skipped."), kind="neutral")
    echo_result
    return (echo_result,)


@app.cell
def _(cfg, flow_params, httpx, mo, tqdm, zipfile):
    def _ingest_sdwis():
        import duckdb as _duckdb
        from pathlib import Path as _P
        bronze_epa = _P(cfg.storage.bronze_path) / "epa"
        bronze_epa.mkdir(parents=True, exist_ok=True)
        dest = bronze_epa / "SDWA_downloads.zip"
        extract_dir = bronze_epa / "sdwis"
        parquet_out = bronze_epa / "sdwis_water_systems.parquet"
        url = cfg.sdwis.sdwa_zip_url

        if not dest.exists():
            try:
                print(f"  Downloading EPA SDWIS from {url}")
                with httpx.stream("GET", url, follow_redirects=True, timeout=300) as r:
                    r.raise_for_status()
                    total = int(r.headers.get("content-length", 0))
                    with open(dest, "wb") as f, tqdm(
                        total=total, unit="B", unit_scale=True, desc="SDWA_downloads.zip", leave=False
                    ) as bar:
                        for chunk in r.iter_bytes(chunk_size=65536):
                            f.write(chunk)
                            bar.update(len(chunk))
                print(f"  EPA SDWIS: downloaded ({dest.stat().st_size / 1_048_576:.1f} MB)")
            except httpx.HTTPStatusError as e:
                dest.unlink(missing_ok=True)
                print(f"  ERROR: EPA SDWIS download failed (HTTP {e.response.status_code})")
                return
            except Exception as e:
                dest.unlink(missing_ok=True)
                print(f"  WARNING: Failed to download EPA SDWIS: {e}")
                return
        else:
            print("  EPA SDWIS: already downloaded, skipping download")

        sentinel = extract_dir / ".extracted"
        if not sentinel.exists():
            extract_dir.mkdir(parents=True, exist_ok=True)
            try:
                with zipfile.ZipFile(dest) as z:
                    print(f"  Extracting {len(z.namelist())} files → {extract_dir}")
                    z.extractall(extract_dir)
                sentinel.touch()
            except zipfile.BadZipFile:
                dest.unlink(missing_ok=True)
                print("  ERROR: SDWA_downloads.zip is corrupt — deleted, re-run to re-download.")
                return
        else:
            print(f"  EPA SDWIS: already extracted → {extract_dir}")

        if parquet_out.exists():
            print(f"  EPA SDWIS Parquet: already exists → {parquet_out}")
            return

        pws_files = (
            sorted(extract_dir.glob("SDWA_PUB_WATER_SYSTEMS*.csv"))
            + sorted(extract_dir.glob("*PUB_WATER*.csv"))
            + sorted(extract_dir.glob("*water_system*.csv"))
        )
        if not pws_files:
            print(f"  WARNING: SDWA_PUB_WATER_SYSTEMS.csv not found in extract.")
            print(f"    Available files: {[f.name for f in extract_dir.iterdir() if f.suffix.lower() == '.csv'][:10]}")
            return

        csv_path = str(pws_files[0]).replace("\\", "/")
        out_path = str(parquet_out).replace("\\", "/")
        print(f"  Converting {pws_files[0].name} → {parquet_out.name}")
        conn = _duckdb.connect()
        try:
            conn.execute("INSTALL spatial; LOAD spatial;")
            # Detect lat/lon column names (may be mixed-case)
            peek = conn.execute(
                f"SELECT * FROM read_csv('{csv_path}', header=True, ignore_errors=True) LIMIT 0"
            ).description
            col_map = {c[0].upper(): c[0] for c in peek}
            lat_col = col_map.get("LATITUDE")
            lon_col = col_map.get("LONGITUDE")
            if lat_col and lon_col:
                conn.execute(f"""
                    COPY (
                        SELECT
                            * EXCLUDE ("{lat_col}", "{lon_col}"),
                            CASE
                                WHEN TRY_CAST("{lat_col}" AS DOUBLE) IS NOT NULL
                                 AND TRY_CAST("{lon_col}" AS DOUBLE) IS NOT NULL
                                THEN ST_Point(
                                    TRY_CAST("{lon_col}" AS DOUBLE),
                                    TRY_CAST("{lat_col}" AS DOUBLE)
                                )
                                ELSE NULL
                            END AS geometry
                        FROM read_csv('{csv_path}', header=True, ignore_errors=True, all_varchar=True)
                    ) TO '{out_path}' (FORMAT PARQUET)
                """)
                geo_note = " (GeoParquet)"
            else:
                conn.execute(f"""
                    COPY (
                        SELECT * FROM read_csv('{csv_path}', header=True, ignore_errors=True)
                    ) TO '{out_path}' (FORMAT PARQUET)
                """)
                geo_note = " (no geometry — lat/lon not in this table)"
            count = conn.execute(f"SELECT COUNT(*) FROM '{out_path}'").fetchone()[0]
            print(f"  EPA SDWIS: {count:,} water systems → {parquet_out.name}{geo_note} "
                  f"({parquet_out.stat().st_size / 1_048_576:.1f} MB)")
        except Exception as e:
            print(f"  ERROR converting SDWIS CSV: {e}")
            parquet_out.unlink(missing_ok=True)
        finally:
            conn.close()

    if flow_params.run_sdwis:
        print("[SDWIS] Starting EPA SDWIS drinking water download + conversion")
        _ingest_sdwis()
        sdwis_result = mo.callout(mo.md("✅ **EPA SDWIS → Parquet complete.**"), kind="success")
    else:
        sdwis_result = mo.callout(mo.md("⏭ EPA SDWIS download skipped."), kind="neutral")
    sdwis_result
    return (sdwis_result,)


@app.cell
def _(cfg, flow_params, httpx, mo, tqdm, zipfile):
    def _ingest_fcc():
        import duckdb as _duckdb
        from pathlib import Path as _P
        bronze_fcc = _P(cfg.storage.bronze_path) / "fcc"
        bronze_fcc.mkdir(parents=True, exist_ok=True)
        dest = bronze_fcc / "r_tower.zip"
        extract_dir = bronze_fcc / "asr"
        parquet_out = bronze_fcc / "asr_towers.parquet"
        url = cfg.fcc.asr_zip_url

        if not dest.exists():
            try:
                print(f"  Downloading FCC ASR from {url}")
                with httpx.stream("GET", url, follow_redirects=True, timeout=600) as r:
                    r.raise_for_status()
                    total = int(r.headers.get("content-length", 0))
                    with open(dest, "wb") as f, tqdm(
                        total=total, unit="B", unit_scale=True, desc="r_tower.zip", leave=False
                    ) as bar:
                        for chunk in r.iter_bytes(chunk_size=65536):
                            f.write(chunk)
                            bar.update(len(chunk))
                print(f"  FCC ASR: downloaded ({dest.stat().st_size / 1_048_576:.1f} MB)")
            except httpx.HTTPStatusError as e:
                dest.unlink(missing_ok=True)
                print(f"  ERROR: FCC ASR download failed (HTTP {e.response.status_code})")
                return
            except Exception as e:
                dest.unlink(missing_ok=True)
                print(f"  WARNING: Failed to download FCC ASR: {e}")
                return
        else:
            print("  FCC ASR: already downloaded, skipping download")

        sentinel = extract_dir / ".extracted"
        if not sentinel.exists():
            extract_dir.mkdir(parents=True, exist_ok=True)
            try:
                with zipfile.ZipFile(dest) as z:
                    print(f"  Extracting {len(z.namelist())} files -> {extract_dir}")
                    z.extractall(extract_dir)
                sentinel.touch()
            except zipfile.BadZipFile:
                dest.unlink(missing_ok=True)
                print("  ERROR: r_tower.zip is corrupt — deleted, re-run to re-download.")
                return
        else:
            print(f"  FCC ASR: already extracted -> {extract_dir}")

        if parquet_out.exists():
            print(f"  FCC ASR GeoParquet: already exists -> {parquet_out}")
            return

        co_file = extract_dir / "CO.dat"
        en_file = extract_dir / "EN.dat"
        if not co_file.exists() or not en_file.exists():
            avail = [f.name for f in extract_dir.iterdir()][:15]
            print(f"  WARNING: Expected CO.dat + EN.dat not found. Files: {avail}")
            return

        co_path  = str(co_file).replace("\\", "/")
        en_path  = str(en_file).replace("\\", "/")
        out_path = str(parquet_out).replace("\\", "/")
        print(f"  Converting CO.dat + EN.dat -> {parquet_out.name}")

        # CO.dat column spec (18 pipe-separated fields, no header)
        # Field positions verified from live data:
        #   2=uls_file_num, 4=call_sign, 9=lat_dir, 10=lat_arcsec, 14=lon_dir, 15=lon_arcsec
        # lat/lon_arcsec = total arc-seconds (deg*3600 + min*60 + sec); divide by 3600 for decimal degrees
        # EN.dat column spec (23 fields):
        #   2=uls_file_num, 8=entity_name, 16=street_address, 19=city, 20=state, 21=zip

        def _col_spec(n):
            return "{" + ", ".join(f"'column{i:02d}': 'VARCHAR'" for i in range(n)) + "}"

        conn = _duckdb.connect()
        try:
            conn.execute("INSTALL spatial; LOAD spatial;")
            conn.execute(f"""
                COPY (
                    WITH co AS (
                        SELECT
                            column02 AS uls_file_num,
                            column04 AS call_sign,
                            column09 AS lat_dir,
                            column10 AS lat_arcsec,
                            column14 AS lon_dir,
                            column15 AS lon_arcsec
                        FROM read_csv('{co_path}', delim='|', header=false,
                            all_varchar=true, ignore_errors=true,
                            columns={_col_spec(18)})
                        WHERE column00 = 'CO'
                    ),
                    en AS (
                        SELECT
                            column02 AS uls_file_num,
                            column09 AS entity_name,
                            column17 AS street_address,
                            column20 AS city,
                            column21 AS state,
                            column22 AS zip
                        FROM read_csv('{en_path}', delim='|', header=false,
                            all_varchar=true, ignore_errors=true, null_padding=true,
                            columns={_col_spec(25)})
                        WHERE column00 = 'EN'
                    ),
                    joined AS (
                        SELECT
                            co.uls_file_num, co.call_sign,
                            en.entity_name, en.street_address, en.city, en.state, en.zip,
                            CASE
                                WHEN TRY_CAST(co.lat_arcsec AS DOUBLE) > 0
                                THEN TRY_CAST(co.lat_arcsec AS DOUBLE) / 3600.0
                                     * CASE WHEN co.lat_dir = 'S' THEN -1.0 ELSE 1.0 END
                            END AS latitude,
                            CASE
                                WHEN TRY_CAST(co.lon_arcsec AS DOUBLE) > 0
                                THEN TRY_CAST(co.lon_arcsec AS DOUBLE) / 3600.0
                                     * CASE WHEN co.lon_dir = 'W' THEN -1.0 ELSE 1.0 END
                            END AS longitude
                        FROM co LEFT JOIN en USING (uls_file_num)
                    )
                    SELECT *,
                        CASE
                            WHEN latitude IS NOT NULL AND longitude IS NOT NULL
                            THEN ST_Point(longitude, latitude)
                        END AS geometry
                    FROM joined
                ) TO '{out_path}' (FORMAT PARQUET, COMPRESSION 'ZSTD')
            """)
            count = conn.execute(f"SELECT COUNT(*) FROM read_parquet('{out_path}')").fetchone()[0]
            with_geom = conn.execute(
                f"SELECT COUNT(*) FROM read_parquet('{out_path}') WHERE geometry IS NOT NULL"
            ).fetchone()[0]
            print(f"  FCC ASR: {count:,} towers ({with_geom:,} with geometry) -> {parquet_out.name} "
                  f"({parquet_out.stat().st_size / 1_048_576:.1f} MB)")
        except Exception as e:
            print(f"  ERROR processing FCC ASR: {e}")
            parquet_out.unlink(missing_ok=True)
        finally:
            conn.close()

    if flow_params.run_fcc:
        print("[FCC] Starting FCC ASR cell tower download + GeoParquet conversion")
        _ingest_fcc()
        fcc_result = mo.callout(mo.md("✅ **FCC ASR → GeoParquet complete.**"), kind="success")
    else:
        fcc_result = mo.callout(mo.md("⏭ FCC ASR download skipped."), kind="neutral")
    fcc_result
    return (fcc_result,)


@app.cell
def _(cfg, flow_params, full_state, httpx, mo, tqdm):
    def _ingest_irs_tcn():
        import duckdb as _duckdb
        import openpyxl as _openpyxl
        from pathlib import Path as _P
        from lib.geocoder import (
            geocode, parse_street_address, normalize_zip,
            GEOCODE_MIN_SCORE,
        )

        bronze_irs = _P(cfg.storage.bronze_path) / "irs"
        bronze_irs.mkdir(parents=True, exist_ok=True)
        dest = bronze_irs / "tcn-db.xlsx"
        parquet_out = bronze_irs / "tcn_terminals.parquet"
        url = cfg.irs.tcn_xlsx_url

        if not dest.exists():
            try:
                print(f"  Downloading IRS TCN database from {url}")
                with httpx.stream("GET", url, follow_redirects=True, timeout=120) as r:
                    r.raise_for_status()
                    total = int(r.headers.get("content-length", 0))
                    with open(dest, "wb") as f, tqdm(
                        total=total, unit="B", unit_scale=True, desc="tcn-db.xlsx", leave=False
                    ) as bar:
                        for chunk in r.iter_bytes(chunk_size=65536):
                            f.write(chunk)
                            bar.update(len(chunk))
                print(f"  IRS TCN: downloaded ({dest.stat().st_size / 1_048_576:.1f} MB)")
            except httpx.HTTPStatusError as e:
                dest.unlink(missing_ok=True)
                print(f"  ERROR: IRS TCN download failed (HTTP {e.response.status_code}). "
                      f"Update irs.tcn_xlsx_url in config.lifeline.yaml.")
                return
            except Exception as e:
                dest.unlink(missing_ok=True)
                print(f"  WARNING: Failed to download IRS TCN: {e}")
                return
        else:
            print("  IRS TCN: already downloaded, skipping download")

        if parquet_out.exists():
            print(f"  IRS TCN GeoParquet: already exists -> {parquet_out}")
            return

        # Read XLSX: row 0 = title, row 1 = headers, row 2+ = data
        print("  Reading tcn-db.xlsx ...")
        wb = _openpyxl.load_workbook(dest, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        # Row 1 = headers (index 1), rows 2+ = data
        if len(rows) < 3:
            print("  ERROR: Unexpected XLSX structure (too few rows).")
            return
        headers = [str(h).strip() if h else f"col{i}" for i, h in enumerate(rows[1])]
        records = []
        for row in rows[2:]:
            if not any(row):
                continue
            records.append(dict(zip(headers, [str(v).strip() if v is not None else "" for v in row])))
        print(f"  Loaded {len(records):,} terminals from XLSX")

        # Geocode against Overture addresses (if available)
        overture_base = str(_P(cfg.storage.bronze_path) / "overture" / "addresses")
        has_overture = _P(overture_base).exists()
        if not has_overture:
            print("  WARNING: Overture address data not found — saving without geometry.")
            print("           Run flows/00_setup.py to enable geocoding.")

        unknown_states: set = set()
        geo_hits = 0
        geo_tried = 0

        geo_conn = _duckdb.connect() if has_overture else None
        if geo_conn:
            geo_conn.execute("INSTALL spatial; LOAD spatial;")

        result_rows = []
        for rec in records:
            termno    = rec.get("TERMNO", "")
            termname  = rec.get("TERMNAME", "")
            addr1     = rec.get("TERMADDR1", "")
            addr2     = rec.get("TERMADDR2", "")
            city      = rec.get("TERMCITY", "")
            state_ab  = rec.get("TERMST", "").strip().upper()
            zip_raw   = rec.get("TERMZIP", "")
            secureair = rec.get("SECUREAIR", "")

            zip5 = normalize_zip(zip_raw)
            # Since Overture 2026-04-15, partitions use 2-letter abbreviations directly
            if state_ab and not state_ab.isalpha():
                unknown_states.add(state_ab)

            housenumber, street = parse_street_address(addr1)
            geocode_wkt = None
            geocode_score = None
            matched_street = None
            matched_city = None
            geocode_status = "skipped"
            geometry_wkb = None

            if has_overture and housenumber and zip5 and state_ab:
                geo_tried += 1
                hits = geocode(
                    street=street, housenumber=housenumber,
                    postcode=zip5, state=state_ab, country="US",
                    base_path=overture_base, limit=3, conn=geo_conn,
                )
                if hits:
                    best = hits[0]
                    score = best["score"]
                    # Validate: score threshold + city cross-check (case-insensitive prefix)
                    city_ok = (
                        city.lower()[:6] in best["postal_city"].lower()
                        or best["postal_city"].lower()[:6] in city.lower()
                    ) if best["postal_city"] else True
                    if score >= GEOCODE_MIN_SCORE and city_ok:
                        geocode_wkt    = best["wkt"]
                        geocode_score  = score
                        matched_street = best["street"]
                        matched_city   = best["postal_city"]
                        geocode_status = "matched"
                        geo_hits += 1
                    else:
                        geocode_status = f"rejected(score={score:.3f},city_ok={city_ok})"
                else:
                    geocode_status = "no_match"
            elif has_overture and not housenumber:
                geocode_status = "no_housenumber"
            elif has_overture and not zip5:
                geocode_status = "no_zip"
            elif has_overture and not full_state:
                geocode_status = f"unknown_state:{state_ab}"

            result_rows.append((
                termno, termname, addr1, addr2, city, state_ab, zip5,
                secureair, geocode_wkt, geocode_score,
                matched_street, matched_city, geocode_status,
            ))

        if geo_conn:
            geo_conn.close()

        if unknown_states:
            print(f"  WARNING: Unrecognized state codes (not geocoded): {sorted(unknown_states)}")

        if has_overture:
            print(f"  Geocoded {geo_hits:,} / {geo_tried:,} tried "
                  f"({geo_hits/geo_tried*100:.1f}% hit rate)"
                  if geo_tried else "  No geocodable records (all lacked housenumber/zip/state)")

        # Write GeoParquet via DuckDB
        out_path = str(parquet_out).replace("\\", "/")
        conn = _duckdb.connect()
        try:
            conn.execute("INSTALL spatial; LOAD spatial;")
            conn.execute("""
                CREATE TABLE tcn (
                    termno VARCHAR, termname VARCHAR, termaddr1 VARCHAR, termaddr2 VARCHAR,
                    termcity VARCHAR, termst VARCHAR, termzip VARCHAR, secureair VARCHAR,
                    geocode_wkt VARCHAR, geocode_score DOUBLE,
                    matched_street VARCHAR, matched_city VARCHAR, geocode_status VARCHAR
                )
            """)
            conn.executemany("INSERT INTO tcn VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)", result_rows)
            conn.execute(f"""
                COPY (
                    SELECT
                        termno, termname, termaddr1, termaddr2,
                        termcity, termst, termzip, secureair,
                        geocode_score, geocode_status,
                        matched_street, matched_city,
                        CASE WHEN geocode_wkt IS NOT NULL
                             THEN ST_GeomFromText(geocode_wkt)
                        END AS geometry
                    FROM tcn
                ) TO '{out_path}' (FORMAT PARQUET, COMPRESSION 'ZSTD')
            """)
            total = conn.execute(f"SELECT COUNT(*) FROM read_parquet('{out_path}')").fetchone()[0]
            with_geom = conn.execute(
                f"SELECT COUNT(*) FROM read_parquet('{out_path}') WHERE geometry IS NOT NULL"
            ).fetchone()[0]
            mb = parquet_out.stat().st_size / 1_048_576
            print(f"  IRS TCN: {total:,} terminals "
                  f"({with_geom:,} geocoded, {total-with_geom:,} without geometry) "
                  f"-> {parquet_out.name} ({mb:.1f} MB)")
        except Exception as e:
            print(f"  ERROR writing IRS TCN GeoParquet: {e}")
            parquet_out.unlink(missing_ok=True)
        finally:
            conn.close()

    if flow_params.run_irs_tcn:
        print("[IRS] Starting IRS TCN fuel terminal download + GeoParquet conversion")
        _ingest_irs_tcn()
        irs_result = mo.callout(mo.md("✅ **IRS TCN → GeoParquet complete.**"), kind="success")
    else:
        irs_result = mo.callout(mo.md("⏭ IRS TCN download skipped."), kind="neutral")
    irs_result
    return (irs_result,)


@app.cell
def _(cfg, flow_params, httpx, mo):
    def _load_env_local():
        """Parse .env.local for key=value pairs (no dependency on python-dotenv)."""
        from pathlib import Path as _P
        env_path = _P(".env.local")
        if not env_path.exists():
            return {}
        result = {}
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                key, _, val = line.partition("=")
                result[key.strip()] = val.strip()
        return result

    def _ingest_eia_terminals():
        import json as _json
        import duckdb as _duckdb
        from pathlib import Path as _P
        env_vars = _load_env_local()
        api_key = env_vars.get(cfg.eia_api.env_key_name, "")
        if not api_key:
            print(f"  INFO: No EIA API key found in .env.local "
                  f"(set {cfg.eia_api.env_key_name}=your_key_here to enable).")
            print("  Register at: https://www.eia.gov/opendata/register.php")
            return None  # signal: skipped, not an error

        bronze_eia = _P(cfg.storage.bronze_path) / "eia"
        bronze_eia.mkdir(parents=True, exist_ok=True)
        parquet_out = bronze_eia / "eia_terminal_stocks.parquet"
        if parquet_out.exists():
            print(f"  EIA Terminals Parquet: already exists → {parquet_out}")
            return True

        # Fetch terminal stocks data from EIA v2 API (pagination via offset)
        base_url = cfg.eia_api.terminal_stocks_url
        all_rows = []
        offset = 0
        length = 5000
        print(f"  Querying EIA API: {base_url}")
        while True:
            params = f"?api_key={api_key}&frequency=monthly&data[0]=value&sort[0][column]=period&sort[0][direction]=desc&offset={offset}&length={length}"
            try:
                resp = httpx.get(base_url + params, follow_redirects=True, timeout=60)
                resp.raise_for_status()
            except httpx.HTTPStatusError as e:
                print(f"  ERROR: EIA API request failed (HTTP {e.response.status_code})")
                return
            except Exception as e:
                print(f"  WARNING: EIA API request failed: {e}")
                return
            data = resp.json().get("response", {})
            rows = data.get("data", [])
            total = data.get("total", 0)
            all_rows.extend(rows)
            offset += length
            if offset >= total or not rows:
                break

        if not all_rows:
            print("  WARNING: EIA API returned no terminal data.")
            return

        # Save via DuckDB
        import tempfile, json as _json2
        tmp_path = parquet_out.with_suffix(".tmp.json")
        tmp_path.write_text(_json2.dumps(all_rows), encoding="utf-8")
        tmp_str = str(tmp_path).replace("\\", "/")
        out_path = str(parquet_out).replace("\\", "/")
        conn = _duckdb.connect()
        try:
            conn.execute(f"COPY (SELECT * FROM read_json('{tmp_str}')) TO '{out_path}' (FORMAT PARQUET)")
            count = conn.execute(f"SELECT COUNT(*) FROM '{out_path}'").fetchone()[0]
            print(f"  EIA Terminals: {count:,} records → {parquet_out.name} "
                  f"({parquet_out.stat().st_size / 1_048_576:.1f} MB) "
                  f"[state/terminal-level data; no GPS coordinates]")
        finally:
            conn.close()
            tmp_path.unlink(missing_ok=True)
        return True

    if flow_params.run_eia_terminals:
        print("[EIA API] Starting EIA terminal data download")
        result = _ingest_eia_terminals()
        if result is None:
            eia_api_result = mo.callout(
                mo.md("ℹ️ **EIA terminals skipped** — set `EIA_API_KEY` in `.env.local` to enable. "
                      "[Register here](https://www.eia.gov/opendata/register.php)"),
                kind="info",
            )
        else:
            eia_api_result = mo.callout(mo.md("✅ **EIA terminal stocks → Parquet complete.**"), kind="success")
    else:
        eia_api_result = mo.callout(mo.md("⏭ EIA terminals download skipped."), kind="neutral")
    eia_api_result
    return (eia_api_result,)


@app.cell
def _(cfg, flow_params, mo):
    if flow_params.run_hifld and cfg.hifld.enabled:
        from pathlib import Path as _P
        from lib.hifld_validation import download_hifld_layer
        bronze_path = _P(cfg.storage.bronze_path)
        print("[HIFLD] Starting HIFLD layer downloads from seerai-hifld-archive")
        _hifld_ok = 0
        _hifld_skip = 0
        for _name, _layer_def in cfg.hifld.layers.items():
            _dest = bronze_path / "hifld" / f"{_name}.parquet"
            print(f"  [{_name}] {_layer_def.gcs_path}")
            try:
                _downloaded = download_hifld_layer(_layer_def.gcs_path, _dest)
                if _downloaded:
                    print(f"    → downloaded ({_dest.stat().st_size / 1_048_576:.1f} MB)")
                    _hifld_ok += 1
                else:
                    print(f"    → already cached, skipping")
                    _hifld_skip += 1
            except Exception as _e:
                print(f"    WARNING: failed to download {_name}: {_e}")
        hifld_result = mo.callout(
            mo.md(f"✅ **HIFLD download complete.** {_hifld_ok} downloaded, {_hifld_skip} cached."),
            kind="success",
        )
    else:
        hifld_result = mo.callout(mo.md("⏭ HIFLD download skipped."), kind="neutral")
    hifld_result
    return (hifld_result,)


@app.cell
def _(cfg, flow_params, mo):
    if flow_params.run_cms and cfg.cms.enabled:
        from pathlib import Path as _P

        _out_dir = _P(cfg.storage.bronze_path) / "cms"
        _out_dir.mkdir(parents=True, exist_ok=True)
        _out_path = _out_dir / "cms_hospital_providers.parquet"

        if _out_path.exists():
            cms_result = mo.callout(
                mo.md("⏭ **CMS providers already cached** — `bronze/cms/cms_hospital_providers.parquet` exists. Delete to re-download."),
                kind="neutral",
            )
        else:
            try:
                from lib.cms_ingest import download_cms_providers as _dl_cms

                _count = _dl_cms(cfg.cms.api_url, cfg.cms.page_size, _out_path)
                _mb = _out_path.stat().st_size / 1_048_576
                cms_result = mo.callout(
                    mo.md(f"✅ **CMS providers downloaded** — {_count:,} records ({_mb:.1f} MB)."),
                    kind="success",
                )
            except Exception as _e:
                print(f"  ERROR: CMS download failed: {_e}")
                cms_result = mo.callout(mo.md(f"⚠️ CMS download failed: {_e}"), kind="warn")
    else:
        cms_result = mo.callout(mo.md("⏭ CMS provider download skipped."), kind="neutral")
    cms_result
    return (cms_result,)


@app.cell
def _(cfg, flow_params, mo, cms_result):  # noqa: F841 — depends on cms_result to run after download
    _geo_path = getattr(cfg.cms, "geocode_address_path", None)
    if (
        flow_params.run_cms
        and cfg.cms.enabled
        and _geo_path
    ):
        from pathlib import Path as _P

        _parquet_path = _P(cfg.storage.bronze_path) / "cms" / "cms_hospital_providers.parquet"
        if not _parquet_path.exists():
            cms_geocode_result = mo.callout(
                mo.md("⏭ CMS geocoding skipped — bronze parquet not found (download may have failed)."),
                kind="neutral",
            )
        elif not _P(_geo_path).exists():
            cms_geocode_result = mo.callout(
                mo.md(f"⚠️ CMS geocoding skipped — `geocode_address_path` not found: `{_geo_path}`"),
                kind="warn",
            )
        else:
            try:
                from lib.cms_ingest import geocode_cms_providers as _geo_cms

                _stats = _geo_cms(
                    _parquet_path,
                    _geo_path,
                    min_score=cfg.cms.geocode_min_score,
                    provider_categories=cfg.cms.geocode_provider_categories,
                )
                _ok = _stats.get("ok", 0)
                _total = sum(_stats.values())
                cms_geocode_result = mo.callout(
                    mo.md(f"✅ **CMS geocoding complete** — {_ok:,}/{_total:,} records geocoded."),
                    kind="success",
                )
            except Exception as _e:
                print(f"  ERROR: CMS geocoding failed: {_e}")
                cms_geocode_result = mo.callout(
                    mo.md(f"⚠️ CMS geocoding failed: {_e}"), kind="warn"
                )
    else:
        cms_geocode_result = mo.callout(
            mo.md("⏭ CMS geocoding skipped — `geocode_address_path` not configured."),
            kind="neutral",
        )
    cms_geocode_result
    return (cms_geocode_result,)


@app.cell
def _(cfg, flow_params, mo, cms_geocode_result):  # noqa: F841 — runs after Overture geocode
    if (
        flow_params.run_cms
        and cfg.cms.enabled
        and getattr(cfg.cms, "geocode_census_fallback", True)
    ):
        from pathlib import Path as _P

        _parquet_path = _P(cfg.storage.bronze_path) / "cms" / "cms_hospital_providers.parquet"
        if not _parquet_path.exists():
            cms_fallback_result = mo.callout(
                mo.md("⏭ CMS fallback geocoding skipped — bronze parquet not found."),
                kind="neutral",
            )
        else:
            try:
                from lib.cms_ingest import fallback_geocode_cms_providers as _fb_cms

                _stats = _fb_cms(
                    _parquet_path,
                    provider_categories=cfg.cms.geocode_provider_categories,
                )
                _ok = _stats.get("ok_census", 0) + _stats.get("ok_nominatim", 0)
                _total = sum(_stats.values())
                _detail = ""
                if _stats.get("ok_census"):
                    _detail += f" Census: {_stats['ok_census']:,}."
                if _stats.get("ok_nominatim"):
                    _detail += f" Nominatim: {_stats['ok_nominatim']:,}."
                cms_fallback_result = mo.callout(
                    mo.md(f"✅ **CMS fallback geocoding complete** — {_ok:,}/{_total:,} additional records geocoded.{_detail}"),
                    kind="success",
                )
            except Exception as _e:
                print(f"  ERROR: CMS fallback geocoding failed: {_e}")
                cms_fallback_result = mo.callout(
                    mo.md(f"⚠️ CMS fallback geocoding failed: {_e}"), kind="warn"
                )
    else:
        cms_fallback_result = mo.callout(
            mo.md("⏭ CMS fallback geocoding skipped."),
            kind="neutral",
        )
    cms_fallback_result
    return (cms_fallback_result,)


@app.cell
def _(
    cms_result,
    cms_geocode_result,
    cms_fallback_result,
    echo_result,
    eia_api_result,
    eia_result,
    epa_result,
    fcc_result,
    hifld_result,
    irs_result,
    mo,
    osm_result,
    sdwis_result,
    is_script_mode,
):
    if is_script_mode:
        print("Ingestion Summary")
        print("Flow 01 complete. Run flows/02_silver_conflation.py next.")
    else:
        _summary = mo.vstack([
            mo.md("## Ingestion Summary"),
            osm_result,
            eia_result,
            epa_result,
            echo_result,
            sdwis_result,
            fcc_result,
            irs_result,
            eia_api_result,
            hifld_result,
            cms_result,
            cms_geocode_result,
            cms_fallback_result,
            mo.callout(mo.md("✅ **Flow 01 complete.** ➡ Run `flows/02_silver_conflation.py` next."), kind="success"),
        ])
        _summary
    return


if __name__ == "__main__":
    app.run()
